#!/usr/bin/env python3
"""
data_init.json generator — 千分の一の国

Reads the statistics files in data/raw/ and writes data/json/data_init.json.

Design rule, from the project's second vow (仮定を全部見せる):
every field carries its own provenance, and provenance is honest about method.

    "parsed"      — extracted from the file named in `source`
    "published"   — a published figure typed in by hand; `source` is where it came from
    "approximated"— not from any source; a modelling assumption. Must say so.

Never label an approximation as parsed. A `sources` block that names a government
statistic for a hand-written number is the one thing that would discredit the whole
simulator, and it is what the previous version of this script did.
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import numpy as np
import openpyxl
import xlrd

ROOT = Path(__file__).resolve().parent
RAW = ROOT / "data" / "raw"
OUT = ROOT / "data" / "json" / "data_init.json"

LIFE_M = RAW / "23rd-lifetable-male.xlsx"
LIFE_F = RAW / "23rd-lifetable-female.xlsx"
INCOME_CSV = RAW / "FEH_00450061_260730122235.csv"
WAGE_XLSX = RAW / "(1-3-1-1)aa1y41.xlsx"
PREF_XLSX = RAW / "【総計】都道府県別人口、人口動態数及び世帯数 26stjin.xlsx"
POP_AGE_JSON = RAW / "population_by_age_estat.json"


# =============================================================================
# 1. Life table — 第23回生命表 (厚労省)
# =============================================================================

def parse_life_table(path: Path) -> tuple[list[int], list[float]]:
    """Age and survival probability lx, normalised so lx[0] == 1.0."""
    ws = openpyxl.load_workbook(path).active
    ages: list[int] = []
    lx: list[float] = []

    for row_idx in range(14, ws.max_row + 1):
        age_cell = ws.cell(row_idx, 2).value
        lx_cell = ws.cell(row_idx, 3).value
        if age_cell is None or lx_cell is None:
            break
        age_str = str(age_cell).replace("年", "").strip()
        if not age_str:
            continue
        try:
            ages.append(int(age_str))
            lx.append(float(lx_cell))
        except (TypeError, ValueError):
            continue

    if not lx:
        raise ValueError(f"no life-table rows parsed from {path.name}")
    base = lx[0]
    return ages, [v / base for v in lx]


# =============================================================================
# 2. Income distribution — 令和6年国民生活基礎調査 第21表
# =============================================================================

def parse_income_distribution() -> dict:
    """世帯数の相対度数分布 by 所得金額階級, 総数 column."""
    rows = list(csv.reader(INCOME_CSV.open(encoding="cp932")))

    COL_MEASURE, COL_BRACKET, COL_TOTAL = 5, 8, 10
    bins: list[str] = []
    share: list[float] = []
    stats: dict[str, float] = {}

    for r in rows:
        if len(r) <= COL_TOTAL:
            continue
        measure = r[COL_MEASURE].strip()
        bracket = r[COL_BRACKET].strip()
        raw = r[COL_TOTAL].strip()
        if not measure or not raw or raw in {"-", "…", "***"}:
            continue

        if measure.startswith("世帯数の相対度数分布"):
            if bracket == "総数":
                continue
            bins.append(bracket)
            share.append(float(raw) / 100.0)
        elif measure.startswith("１世帯当たり平均所得金額"):
            stats["meanManYen"] = float(raw)
        elif measure.startswith("中央値"):
            stats["medianManYen"] = float(raw)
        elif measure.startswith("平均所得金額以下の世帯の割合"):
            stats["shareBelowMean"] = float(raw) / 100.0

    if not bins:
        raise ValueError("no income brackets parsed")

    total = sum(share)
    if not 0.98 <= total <= 1.02:
        raise ValueError(f"income shares sum to {total:.4f}, expected ~1.0")
    share = [s / total for s in share]

    return {
        "binsManYen": bins,
        "share": share,
        "edgesManYen": [bracket_edges(b) for b in bins],
        **stats,
    }


def bracket_edges(label: str) -> list[float | None]:
    """'３００～３５０万円未満' -> [300, 350]; '２０００万円以上' -> [2000, None]."""
    z = label.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    nums = [float(n) for n in re.findall(r"\d+", z)]
    if "以上" in z and len(nums) == 1:
        return [nums[0], None]
    if "未満" in z and len(nums) == 1:
        return [0.0, nums[0]]
    if len(nums) >= 2:
        return [nums[0], nums[1]]
    return [None, None]


# =============================================================================
# 3. Wage by age — 令和7年賃金構造基本統計調査 役職第1表
# =============================================================================

def parse_wage_by_age() -> dict:
    """所定内給与額 by age band. File is in 千円; converted to 万円/月 here."""
    ws = openpyxl.load_workbook(WAGE_XLSX, read_only=True)["産業計(役職計)"]

    COL_LABEL, COL_WAGE_SENYEN = 2, 8
    out: dict[str, float] = {}

    # The sheet repeats the whole age ladder once per 学歴 (学歴計, then 中学, 高校 …).
    # Only the first block — 男女計・学歴計 — is wanted, so stop at the first label
    # that isn't an age band. Reading further silently overwrites good values with
    # 中学 figures (20~24 becomes 275.2 instead of 294.1).
    for row in ws.iter_rows(min_row=14, max_row=60, values_only=True):
        if len(row) <= COL_WAGE_SENYEN:
            continue
        label = str(row[COL_LABEL] or "").strip().replace("\n", "")
        if not label:
            continue
        z = label.translate(str.maketrans("０１２３４５６７８９～", "0123456789~"))
        z = z.replace("歳", "").strip()
        if not re.match(r"^(~?\d+|\d+~\d*)$", z):
            if out:
                break          # end of the 学歴計 block
            continue
        wage = row[COL_WAGE_SENYEN]
        if wage is None:
            continue
        try:
            out[z] = round(float(wage) / 10.0, 2)   # 千円 -> 万円
        except (TypeError, ValueError):
            continue

    if not out:
        raise ValueError("no wage bands parsed")
    return out


# =============================================================================
# 4. National population totals — 令和8年1月1日 住民基本台帳
# =============================================================================

def parse_national_totals() -> dict:
    ws = openpyxl.load_workbook(PREF_XLSX, read_only=True).active
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if row and str(row[1]).strip() == "合計":
            male, female, total = int(row[2]), int(row[3]), int(row[4])
            return {
                "male": male,
                "female": female,
                "total": total,
                "households": int(row[5]),
                "femalePerMale": round(female / male, 5),
            }
    raise ValueError("合計 row not found in prefecture file")


# =============================================================================
# 5. Age structure — 人口推計 各年10月1日現在人口 統計表001 (総務省統計局, e-Stat API)
# =============================================================================

def parse_age_structure() -> tuple[dict, str]:
    """
    年齢(各歳)，男女別人口 — via e-Stat API (tools/estat_fetch.py), statsDataId 0003448228.

    cat01: 001=男女計 002=男 003=女 004=人口性比
    cat02: 001=総人口 002=日本人人口
    cat03: 01000=総数, 01001=0歳 .. 01100=99歳, 01101=100歳以上
    time : latest of 5 available years (2020-2024年10月1日現在)
    unit : 千人
    """
    if not POP_AGE_JSON.exists():
        raise FileNotFoundError(
            f"{POP_AGE_JSON.name} not found. Run: "
            f".venv/bin/python tools/estat_fetch.py"
        )
    body = json.loads(POP_AGE_JSON.read_text(encoding="utf-8"))
    data = body["GET_STATS_DATA"]["STATISTICAL_DATA"]
    values = data["DATA_INF"]["VALUE"]

    latest_time = max(v["@time"] for v in values)
    time_obj = next(
        c for c in data["CLASS_INF"]["CLASS_OBJ"] if c["@id"] == "time"
    )["CLASS"]
    time_label = next(c["@name"] for c in time_obj if c["@code"] == latest_time)

    by_sex_age: dict[str, dict[int, float]] = {"002": {}, "003": {}}
    for v in values:
        if v["@time"] != latest_time or v["@cat02"] != "001" or v["@cat01"] not in by_sex_age:
            continue
        code = v["@cat03"]
        if code == "01000":
            continue  # 総数, not a single age
        age = 100 if code == "01101" else int(code[2:]) - 1  # 01001 -> 0歳
        by_sex_age[v["@cat01"]][age] = float(v["$"])  # 千人

    ages = list(range(0, 111))
    male = [by_sex_age["002"].get(a, 0.0) for a in ages]
    female = [by_sex_age["003"].get(a, 0.0) for a in ages]
    if not any(male) or not any(female):
        raise ValueError("population_by_age_estat.json parsed to all zeros — check category codes")

    total = sum(male) + sum(female)
    return {
        "ages": ages,
        "male": [round(m / total * 100, 6) for m in male],
        "female": [round(f / total * 100, 6) for f in female],
    }, time_label


# =============================================================================
# 6. Published figures typed by hand — honest about being hand-entered
# =============================================================================

SPEND_SHARE = {
    "食料": 0.26, "住居": 0.15, "光熱水道": 0.07, "交通通信": 0.13,
    "教育": 0.03, "教養娯楽": 0.09, "保健医療": 0.06, "その他": 0.21,
}

HOUSEHOLD_TYPES = {
    "単独": 0.38, "夫婦のみ": 0.25, "夫婦+子": 0.25, "ひとり親+子": 0.09, "その他": 0.03,
}


# =============================================================================
# Main
# =============================================================================

def main() -> dict:
    print("building data_init.json")

    male_ages, male_lx = parse_life_table(LIFE_M)
    female_ages, female_lx = parse_life_table(LIFE_F)
    print(f"  life table   : male {len(male_ages)} ages, female {len(female_ages)} ages  [parsed]")

    income = parse_income_distribution()
    print(f"  income       : {len(income['binsManYen'])} brackets, "
          f"mean {income.get('meanManYen')}万円, median {income.get('medianManYen')}万円  [parsed]")

    wage = parse_wage_by_age()
    print(f"  wage by age  : {len(wage)} bands, 20-24 = {wage.get('20~24')}万円/月  [parsed]")

    totals = parse_national_totals()
    print(f"  national     : {totals['total']:,} people, "
          f"female/male {totals['femalePerMale']}  [parsed]")

    population, pop_time_label = parse_age_structure()
    print(f"  age structure: {len(population['ages'])} ages, {pop_time_label}  [parsed]")

    init = {
        "population": population,
        "nationalTotals": totals,
        "lifeTable": {
            "male": {"ages": male_ages, "survival": male_lx},
            "female": {"ages": female_ages, "survival": female_lx},
        },
        "incomeHistogram": income,
        "wageByAge": wage,
        "wageUnit": "万円/月 (所定内給与額)",
        "spendShare": SPEND_SHARE,
        "householdTypes": HOUSEHOLD_TYPES,
        "provenance": {
            "lifeTable": {
                "method": "parsed", "source": "第23回生命表(厚生労働省)",
                "file": LIFE_M.name + " / " + LIFE_F.name,
            },
            "incomeHistogram": {
                "method": "parsed", "source": "令和6年国民生活基礎調査 所得 第21表(厚生労働省)",
                "file": INCOME_CSV.name,
                "note": "世帯数の相対度数分布, 世帯類型=総数",
            },
            "wageByAge": {
                "method": "parsed", "source": "令和7年賃金構造基本統計調査 役職第1表(厚生労働省)",
                "file": WAGE_XLSX.name,
                "note": "所定内給与額, 産業計(役職計), 男女計・学歴計. 千円→万円に変換",
            },
            "nationalTotals": {
                "method": "parsed", "source": "令和8年1月1日 住民基本台帳人口・世帯数(総務省)",
                "file": PREF_XLSX.name,
            },
            "population": {
                "method": "parsed",
                "source": "人口推計 各年10月1日現在人口(令和2年国勢調査基準) 統計表001"
                           "(総務省統計局, e-Stat API)",
                "file": POP_AGE_JSON.name,
                "note": f"{pop_time_label}, 総人口, 年齢各歳(0〜99歳)+100歳以上, "
                        f"statsDataId=0003448228",
            },
            "spendShare": {
                "method": "published", "source": "家計調査(総務省統計局)",
                "note": "公表値を手入力。ファイルからの抽出ではない。",
            },
            "householdTypes": {
                "method": "published", "source": "国民生活基礎調査(厚生労働省)",
                "note": "公表値を手入力。ファイルからの抽出ではない。",
            },
        },
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(init, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\nwrote {OUT.relative_to(ROOT)}")
    approx = [k for k, v in init["provenance"].items() if v["method"] == "approximated"]
    if approx:
        print(f"STILL APPROXIMATED: {', '.join(approx)} — download the source and re-run")
    return init


if __name__ == "__main__":
    main()
